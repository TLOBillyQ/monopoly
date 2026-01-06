#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import json
from pathlib import Path

docs_dir = Path("docs")

def read_excel_to_dict(file_path):
    """Read Excel file and convert to dictionary"""
    wb = openpyxl.load_workbook(file_path)
    result = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        headers = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:
                headers = [h for h in row if h is not None]
            else:
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i]
                    data.append(row_dict)
        
        result[sheet_name] = data
    
    return result

# Read all Excel files
files = {
    "常量表": "蛋仔--大富翁--常量表.xlsx",
    "角色表": "蛋仔--大富翁--角色表.xlsx",
    "地块表": "蛋仔--大富翁--地块表.xlsx",
    "机会表": "蛋仔--大富翁--机会表.xlsx",
    "道具表": "蛋仔--大富翁--道具表.xlsx",
    "座驾表": "蛋仔--大富翁--座驾表.xlsx",
}

for name, filename in files.items():
    file_path = docs_dir / filename
    if file_path.exists():
        print(f"\n{'='*60}")
        print(f"📄 {name}: {filename}")
        print('='*60)
        data = read_excel_to_dict(file_path)
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        print(f"❌ 文件不存在: {file_path}")
